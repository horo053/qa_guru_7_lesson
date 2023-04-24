import os.path
import csv
import zipfile

from pypdf import PdfReader
import xlrd
from openpyxl import load_workbook
import time
from selenium import webdriver
from selene import browser
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import requests


def test_csv():
# TODO оформить в тест, добавить ассерты и использовать универсальный путь
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    path_csv = os.path.join(PROJECT_ROOT_PATH, 'resources', 'eggs.csv')
    with open(path_csv, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(path_csv) as csvfile:
        csvreader = csv.reader(csvfile)
        list_csv = []
        for row in csvreader:
            list_csv.append(row)
            print(row)

    assert list_csv[0] == ['Anna', 'Pavel', 'Peter']


def test_pdf():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    path_pdf = os.path.join(PROJECT_ROOT_PATH, 'resources', 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(path_pdf)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print(page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412


def test_xls():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    path_xls = os.path.join(PROJECT_ROOT_PATH, 'resources', 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(path_xls)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))

    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']
    assert sheet.ncols == 8
    assert sheet.nrows == 10
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'


def test_xlsx():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    path_xlsx = os.path.join(PROJECT_ROOT_PATH, 'resources', 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(path_xlsx)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'


def test_download_file_with_browser():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    path_browser = os.path.join(PROJECT_ROOT_PATH, 'resources')

    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": path_browser,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(10)

    assert os.path.exists(os.path.join(path_browser, 'pytest-main.zip'))
    assert os.path.getsize(os.path.join(path_browser, 'pytest-main.zip')) > 0


def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    path_file = os.path.join(PROJECT_ROOT_PATH, 'resources')
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'

    r = requests.get(url)
    downloaded_file_path = os.path.join(path_file, 'selenium_logo.png')

    with open(downloaded_file_path, 'wb') as file:
        file.write(r.content)

    size = os.path.getsize(downloaded_file_path)

    assert size == 30803


def test_zip():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    path_dir = os.path.join(PROJECT_ROOT_PATH, 'resources')
    path_zip = os.path.join(PROJECT_ROOT_PATH, 'resources/archive.zip')
    file_zip = zipfile.ZipFile(path_zip, 'w')
    file_list = ['archive.zip', 'docs-pytest-org-en-latest.pdf', 'eggs.csv', 'file_example_XLSX_50.xlsx', 'file_example_XLS_10.xls', 'selenium_logo.png']
    for folder, subfolders, files in os.walk(path_dir):
        for file in files:
            file_zip.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder, file), 'resources'), compress_type=zipfile.ZIP_DEFLATED)
    file_zip.close()

    l = []
    with zipfile.ZipFile(path_zip, mode='a') as zf:
        for file in zf.namelist():
            l.append(file)

    assert l == file_list